############################################################################################################################################################
############################################################################################################################################################ 

# Program name: abakus_class.py
# Author: Luca Teruzzi, PhD in Physics, Astrophysics and Applied Physics, University of Milan, Milan (IT)
#         EuroCold Lab, Department of Earth and Environmental Sciences, University of Milano-Bicocca, Milan (IT)
# Date: 01 May 2022 (last modified)
# Objective: Class definition: Abakus laser sensor
# Description: This Python3 script defines a class named 'Abakus', which allows the user to perform several operations with the Abakus laser sensor provided by
#              Klotz GmbH. 
#              First, is the initialization of the class instance, provinding the serial port connected the the instrument. 
#              Some functions can also be used to communicate via serial port with the Abkaus laser sensor, reading and writing strings ad lines with a specific format 
#              and to check if the instrument is correctly connected to the specified port.
#              At the same time, other functions allows to decode (convert the serial string into something more familiar and user friendly) the serial command 
#              read from the Abakus and to perform some post-processing: retrieve sample concentration, the number of total particles detected, cumulative distributions 
#              and histograms.
#              Each function is fully commented and explained.

############################################################################################################################################################
############################################################################################################################################################


import serial, pandas as pd, numpy as np, os, time                                                      # Import the required libraries
from termcolor import colored
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color


############################################################################################################################################################
############################################################################################################################################################


class Abakus(object):


    # -----------------------------------------------------------------------------------------------------------------------------------------------------#
    # Class constructor: creates an Abakus object.

    def __init__(self, port, baudrate = 38400, bytesize = 8, parity = 'N', stopbits = 1, timeout = 0.1, debug = True): 
        self.port = port                                                                                # Serial port 
        self.debug = debug                                                                              # Debug option (printing results)
        if port!='_default':
            self._dev = serial.Serial(port, baudrate, bytesize, parity, stopbits, timeout)              # Creation of a Abakus object with the specificed serial
                                                                                                        # communication parameters: baudrate, timeout, parity, stopbits and bytesize

    # -----------------------------------------------------------------------------------------------------------------------------------------------------#
    # Function to send serial commands to the Abakus laser sensor.

    def sendCommand(self, cmd):                                                                         # Send string (bytes) to the serial port
        self._dev.write(cmd)

    
    # -----------------------------------------------------------------------------------------------------------------------------------------------------#
    # Method for the definition of some paramteres later implied in the measurement and in the subsequent analysis.

    def abakus_parameters(self, ID_number, path, filename, skiprows, flow_rate, sizes, acquisition_time, delay, time_str, save_path, txt_save_name, model, material, range, terminal, msg_window, err_window):

        self.path = path                                                                                # Directory path
        self.filename = filename                                                                        # File name
        self.skiprows = skiprows                                                                        # NUmber of heading lines to skip
        self.flow_rate = flow_rate                                                                      # Flow rate
        self.sizes = sizes                                                                              # Sizes range during Abakus measurement
        self.save_path = save_path                                                                      # Save path
        self.txt_save_name = txt_save_name                                                              # Output text file name
        self.acquisition_time = acquisition_time                                                        # Acquisition time
        self.terminal = terminal                                                                        # Boolean option for debugging on user command line
        self.err_window =  err_window                                                                   # Errors and warnings output window
        self.window =  msg_window                                                                       # Messages output window
        self.time_delay = delay                                                                         # Time dealy for serial writing/reading
        self.time_str = time_str                                                                        # Date and time 
        self.ID_number = ID_number                                                                      # Abakus ID number
        self.model = model                                                                              # Abakus model
        self.cell_material = material                                                                   # Inner cell material
        self.size_range = range                                                                         # Default Abakus measurement size range


    # -----------------------------------------------------------------------------------------------------------------------------------------------------#
    # Initialization of the text file for saving Abakus measurements.
    # The heading lines of this file specify the kind of software, the flow rate set for the measurement and the noise level, previously calculated via serial 
    # communication with the Abakus laser sensor.
    # Also, the date and time are saved in the file and some checks allow to print in the dedicated GUI window error and warning messages.
    
    def starting_files(self, flow_rate, port, software, noise_level, size_channels, xcell, ycell, zcell_laser, wavelength):

        self.temp = datetime.now().strftime("%d-%m-%Y_%H-%M-%S.%f")[11:-7]

        self.full_path = f"{self.path}{self.time_str[:-12]}/"

        if os.path.isdir(self.full_path): self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t Saving data in '+"'"+str(self.full_path)+"'")
        else: 
            os.makedirs(self.full_path)
            self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t Saving data in '+"'"+str(self.full_path)+"'")
        if self.txt_save_name=='': self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t WARNING: file name not specified. Please fill the dedicate dspace as (filename).txt')
        if software=='': self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t WARNING: Missing software informations.')
        if noise_level==[]: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR: Missing noise levels for Abakus laser sensor measurement!')

        txt_file = open(self.full_path+self.txt_save_name+'_'+self.temp+'.txt', 'w', encoding='utf-8')

        self.work_book = Workbook()
        self.initial_worksheet = self.work_book.active
        self.initial_worksheet = self.work_book.create_sheet('data', 0)
        self.initial_worksheet.sheet_properties.tabColor = 'FF0000'
        self.initial_worksheet.column_dimensions['A'].width = '43'
        self.initial_worksheet.column_dimensions['B'].width = '32'
        self.initial_worksheet.column_dimensions['C'].width = '25'
        self.initial_worksheet.column_dimensions['D'].width = '25'

        txt_file.write('ABAKUS LASER SENSOR ----- PARTICLE SIZE DISTRIBUTION DATA\n')                   # File title
        txt_file.write('\nSerial port connected:\t\t\t\t\t\t\t'+str(port))                              # Serial port connected (COM#, /dev/tty#)
        txt_file.write('\nAbakus sofware version:\t\t\t\t\t\t\t'+str(software))                         # Software default version
        txt_file.write('\nAbakus model:\t\t\t\t\t\t\t\t\t'+self.model)                                  # ABakus mmodel
        txt_file.write('\nAbakus ID number:\t\t\t\t\t\t\t\t'+self.ID_number)                            # Abakus ID number
        txt_file.write('\nAbakus cell dimensions:\t\t\t\t\t\t\t'+str(xcell)+' x '+str(ycell)+' μm^2')   # Abakus cell dimensions and material
        txt_file.write('\nAbakus cell material:\t\t\t\t\t\t\t'+self.cell_material)
        txt_file.write('\nAbakus laser wavelength:\t\t\t\t\t\t'+'{:.03f}'.format(wavelength)+' μm')     # Wavelength
        txt_file.write('\nAbakus laser waist:\t\t\t\t\t\t\t\t'+str(zcell_laser)+' μm')                  # Abakus laser waist
        txt_file.write('\nDetectable size range:\t\t\t\t\t\t\t'+self.size_range+' μm')                  # Abakus measuring size range
        txt_file.write('\nNoise levels and calibration:\n')                                             # Voltage signal on each of the 32 instrument channels [mV]
        for j in range(0, len(noise_level), 2):
            if j<10: 
                if j+2!=10: txt_file.write('\t\t\t\t\t\t\t\t\t\t\t\t'+str((j+1)//2 + 1)+') '+str(noise_level[j+1])+' μm\t--->\t'+str(10*noise_level[j])+'\n')
                else: txt_file.write('\t\t\t\t\t\t\t\t\t\t\t\t'+str((j+1)//2 + 1)+') '+str(noise_level[j+1])+' μm\t--->\t'+str(10*noise_level[j])+'\n')
            elif j>=10: txt_file.write('\t\t\t\t\t\t\t\t\t\t\t\t'+str((j+1)//2 + 1)+') '+str(noise_level[j+1])+' μm\t--->\t'+str(10*noise_level[j])+'\n')
        txt_file.write('Delay time between serial writing and reading:\t'+str(self.time_delay)+' ms')   # Delay time between serial writing and reading (settable by the user)
        txt_file.write('\nFlow rate:\t\t\t\t\t\t\t\t\t\t'+str(flow_rate)+' mL/min')                     # Flow rate
        txt_file.write('\nDate and starting time:\t\t\t\t\t\t\t'+self.time_str+'\n\n')                  # Date and time of acquisition

        txt_file.write('\n_________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________')
        txt_file.write('\n\nIndex\tDuration [s]\tLaser diode voltage[mV]\t\tRAM-buffer voltage [mV]\t\t')
        for i in range(0, len(size_channels)): txt_file.write(str(size_channels[i])+'\t\t')
        txt_file.write('\n_________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________\n')

        self.initial_worksheet.append(['ABAKUS LASER SENSOR ----- PARTICLE SIZE DISTRIBUTION DATA'])            # File title
        self.initial_worksheet.append([''])
        self.initial_worksheet.append(['Serial port connected:', str(port)])                                    # Serial port connected (COM#, /dev/tty#)
        self.initial_worksheet.append(['Abakus sofware version:', str(software)])                               # Software default version
        self.initial_worksheet.append(['Abakus model:', self.model])                                            # ABakus mmodel
        self.initial_worksheet.append(['Abakus ID number:', self.ID_number])                                    # Abakus ID number
        self.initial_worksheet.append(['Abakus cell dimensions:', str(xcell)+' x '+str(ycell)+' μm^2'])         # Abakus cell dimensions and material
        self.initial_worksheet.append(['Abakus cell material:', self.cell_material])
        self.initial_worksheet.append(['Abakus laser wavelength:', '{:.03f}'.format(wavelength)+' μm'])         # Wavelength
        self.initial_worksheet.append(['Abakus laser waist:', str(zcell_laser)+' μm'])                          # Abakus laser waist
        self.initial_worksheet.append(['Detectable size range:', self.size_range+' μm'])                        # Abakus measuring size range
        self.initial_worksheet.append(['Noise levels and calibration:'])                                        # Voltage signal on each of the 32 instrument channels [mV]
        for j in range(0, len(noise_level), 2):
            if j<10: 
                if j+2!=10: self.initial_worksheet.append(['', str((j+1)//2 + 1)+') '+str(noise_level[j+1])+' μm    ------->', str(10*noise_level[j])+' mV'])
                else: self.initial_worksheet.append(['', str((j+1)//2 + 1)+') '+str(noise_level[j+1])+' μm    ------->', str(10*noise_level[j])+' mV'])
            elif j>=10: self.initial_worksheet.append(['', str((j+1)//2 + 1)+') '+str(noise_level[j+1])+' μm    ------->', str(10*noise_level[j])+' mV'])
        self.initial_worksheet.append([''])
        self.initial_worksheet.append(['Delay time between serial writing and reading:', str(self.time_delay), 'ms'])   # Delay time between serial writing and reading (settable by the user)
        self.initial_worksheet.append(['Starting flow rate:', str(flow_rate), 'mL/min'])                        # Flow rate
        self.initial_worksheet.append(['Date and starting time:', self.time_str])                               # Date and time of acquisition

        self.initial_worksheet.append([''])
        self.initial_worksheet.append([''])
        self.xlsx_size_channels_list = []
        for i in range(0, len(size_channels)): self.xlsx_size_channels_list.append(str(size_channels[i]))
        self.initial_worksheet.append(['Index', 'Duration [s]', 'Laser diode voltage[mV]', 'RAM-buffer voltage [mV]']+self.xlsx_size_channels_list)

        return txt_file, self.initial_worksheet, self.work_book, self.full_path, self.txt_save_name, self.temp


    # -----------------------------------------------------------------------------------------------------------------------------------------------------#
    # Load and read the content of an Abakus text file, skipping the first #'skip' heading rows; data are organized in a dataframe using pandas library.

    def read_file(self, path, filename, skip):
        read_file = open(path+filename, 'r', encoding='utf-8')                                          # Import text file, extract the flow rate and convert the rest to a data frame
        self.abakus_dataframe = pd.read_csv(read_file, sep="\t", skiprows=skip, header=None)

        read_file1 = open(path+filename, 'r')
        lines = read_file1.readlines()
        self.measd_flow_rate = float(lines[29][-11:-8])                                                 # Flow rate value
        self.noises = lines[12:28]                                                                      # Voltages values


    # -----------------------------------------------------------------------------------------------------------------------------------------------------#
    # Method for sending a command via serial port to the Abakus laser sensor and read the answer after the specified time sleep (tipically measured in ms).
    # If the instrument output is empty, the method waits Δt = time sleep [ms] ms and checks one more time.

    def serial_write_and_read(self, command):

        if command == b'C0001\n': time_sleep = 0.7                                                      # Serial command b'C0001' is sent to start remote control mode
        else: time_sleep = int(self.time_delay)/1000                                                    # after this command is sent, the system waits 0.7 s, otherwise
                                                                                                        # for any other serial command the time gap is set to int(self.time_delay)/1000 [ms]
        self._dev.write(command)                                                                        # Serial writing
        
        time.sleep(time_sleep)                                                                          # Time gap
        answer = self._dev.readline()                                                                   # Serial reading
        if answer==b'':                                                                                 # If the answer is empty, wait until something comes up
            while answer==b'':
                self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t Waiting for Abakus answer via serial port...')
                time.sleep(time_sleep)
                answer = self._dev.readline()
        
        time.sleep(time_sleep)                                                                          # Time gap °2

        return answer


    # -----------------------------------------------------------------------------------------------------------------------------------------------------#
    # Method to unpack and convert the serial output relative to different commands that can be send to the Abakus laser sensor.
    # The main available commands are listed in the following: {'C0001', 'C0003', 'C0004', 'C0013', 'C0012'} for getting the size bins and the corresponding
    # number of counts, {'U0003', 'U0004'} for getting the laser diode and the RAM-buffer intrument voltages, 'X0003' for retrieving the Abakus model and  
    # software version.

    def convert(self, bytes_answer):

        self.string_answer = bytes_answer.decode('utf-8').split(' ')                                    # Decode the bytestring obtained as serial answer
        header_cmd = self.string_answer[0]                                                              # Get the first element of the answer string 

        if self.terminal==True: print('\n', header_cmd, self.string_answer)

        if header_cmd=='C0001' or header_cmd=='C0003' or header_cmd=='C0004' or header_cmd=='C0013':    # Get sizes and number of measured counts
            data = np.zeros(len(self.sizes))
            for k in range(0, len(data)-1):                                                             # If the input command was 'C0001', 'C0003', 'C0004' or 'C0013'
                try: data[k] = float(self.string_answer[k+1])/10                                        # each element in the answer string is conferted to a float value
                except: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR ('+header_cmd+'): element n.'+str(k)+' in serial answer cannot be converted to float.')

        if header_cmd=='C0012':                                                                         # Same as before (different format): if the input command was 'C0012', then the 
            data = np.zeros(2*len(self.sizes))                                                          # answer string is split in two different ways.
            for k in range(0, len(data)-1, 2):                                                          # --> See the Abakus serial communication protocol for further information
                try: data[k] = float(self.string_answer[k+1])/10
                except: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR ('+header_cmd+'): element n.'+str(k)+' in serial answer cannot be converted to float.')
            for kk in range(1, len(data)-1, 2): 
                try: data[kk] = float(self.string_answer[kk+1])
                except: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR ('+header_cmd+'): element n.'+str(kk)+' in serial answer cannot be converted to float.')

        elif header_cmd=='U0003' or header_cmd[-5:]=='U0004':                                           # Get Abakus voltages (both RAM-buffer voltage and regulating laser voltage)
            data = 0
            try: data = float(self.string_answer[1][:-1])
            except: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR ('+header_cmd+'): element n.'+str(k)+' in serial answer cannot be converted to float.')

        elif header_cmd=='X0003':                                                                       # Get the Abakus mmodel and software version
            data = []
            if len(self.string_answer)>1: data = [''.join(self.string_answer[1:])]

        return header_cmd, data


    # -----------------------------------------------------------------------------------------------------------------------------------------------------#
    # This method initialize the Abakus laser sensor and get the basic informations: number of channels of the instrument, noise level for each channel and the
    # software version.

    def initialization(self, channel_cmd, software_cmd, noise_cmd):

        self.channels, self.software, self.noise = '', '', ''                                           # Empy initialization

        try: self.channels = self.convert(self.serial_write_and_read(channel_cmd))                      # Get and convert the measuring channels (aka: the size range)
        except: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR: Can not get channels infromation from Abakus. Please check the input string.')
        self.window.append('Command '+channel_cmd.decode('utf-8')[:-1]+' sent to Abakus: listing Abakus channels.')

        try: self.software = self.convert(self.serial_write_and_read(software_cmd))                     # Get and convert the software version
        except: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR: Can not get software infromation from Abakus. Please check the input string.')
        self.window.append('Command '+software_cmd.decode('utf-8')[:-1]+' sent to Abakus: getting Abakus software version.')

        try: self.noise = self.convert(self.serial_write_and_read(noise_cmd))                           # Get and convert the voltage noise per channel
        except: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR: Can not get noise levels infromation from Abakus. Please check the input string.')
        self.window.append('Command '+noise_cmd.decode('utf-8')[:-1]+' sent to Abakus: measuring noise level for each Abakus channel.')

        if self.channels[0]!=channel_cmd.decode('utf-8')[:-1]: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR('+channel_cmd+'): Input and output commands do not match.')
        if self.software[0]!=software_cmd.decode('utf-8')[:-1]: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR('+software_cmd+'): Input and output commands do not match.')
        if self.noise[0]!=noise_cmd.decode('utf-8')[:-1]: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR('+noise_cmd+'): Input and output commands do not match.')

        return self.channels, self.software, self.noise


    # -----------------------------------------------------------------------------------------------------------------------------------------------------#
    # Method to perform the whole (single) measurement with the Abakus laser sensor.

    def single_measurement(self, data_cmd, volt_cmd, buffer_cmd):

        init_time = datetime.now()                                                                      # Starting time
        self.buffer, self.volt, self.meas_data, self.running_label = '', '', '', False                  # Empy initialization

        if self.terminal==True: print(self.buffer, self.volt, self.meas_data, self.running_label)       # If True, print values and results on terminal
        if self.terminal==True: print(data_cmd, volt_cmd, buffer_cmd)
        if self.terminal==True: print(data_cmd.decode('utf-8')[:-1], volt_cmd.decode('utf-8')[:-1], buffer_cmd.decode('utf-8')[:-1])

        try: self.volt = self.convert(self.serial_write_and_read(volt_cmd))                             # Get laser regulating voltage
        except: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR: Can not get voltage infromation from Abakus. Please check the input string.')

        try: self.buffer = self.convert(self.serial_write_and_read(buffer_cmd))                         # Get RAM-buffer voltage
        except: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR: Can not get RAM-buffer voltage infromation from Abakus. Please check the input string.')
        
        try: 
            self.meas_data = self.convert(self.serial_write_and_read(data_cmd))                         # Get number of counts
            self.running_label = True
        except: self.err_window.append(datetime.now().strftime("%d-%m-%Y_%H:%M:%S.%f")[11:-7]+'\t ERROR: Can not retrieve countings infromation from Abakus. Please check the input string.')
        
        if self.terminal==True: print(self.buffer, self.volt, self.meas_data, self.running_label)

        end_time = datetime.now()                                                                       # Ending time

        return self.volt[1], self.buffer[1], self.meas_data[1], init_time, end_time, self.running_label


    # -----------------------------------------------------------------------------------------------------------------------------------------------------#
    # Method to elaborate the dataframe from 'read_file()' specifying only the rows and columns of interest for the subsequnt analysis; thus, a subset of the original 
    # dataframe is extracted. The total particle concentration is calculated as well as the partial concentration for each instrument channel 
    # (aka: for each particle diameter detected).
    # Lastly, also the cumulative distribution is computed (and plotted as boxplot and as a line in the main program).

    def channel_counts(self, path, filename, skip, sizes):

        self.read_file(path, filename, skip)                                                            # Creating a dataframe from .txt file

        startrow, startcolumn = 0, 16                                                                   # Select rows and columns of interest in dataframe
        endrow, endcolumn = self.abakus_dataframe.shape[0], self.abakus_dataframe.shape[1]
        flow_rate = self.measd_flow_rate/60                                                             # Unit conversion min -> s
        volume = flow_rate*(endrow-startrow)                                                            # Volume set on peristaltic pump [mL]

        self.volt_list = self.abakus_dataframe.loc[startrow:endrow, 6]                                  # Get the laser diode and RAM-buffer voltages from the dataframe
        self.RAM_list = self.abakus_dataframe.loc[startrow:endrow, 12]

        abakus_subset = self.abakus_dataframe.loc[startrow:endrow, startcolumn:endcolumn]
        for i in range(startcolumn, endcolumn-1, 2): abakus_subset.rename(columns={i: sizes[(i-startcolumn)//2]}, inplace=True)
        abakus_subset = abakus_subset[abakus_subset.columns[~abakus_subset.isnull().all()]]

        ptc_concentration = sum(abakus_subset.sum(axis=0)/volume)                                       # Total particles concentration [pt/mL]
        ptc_concentration_sizelist = []
        if self.terminal==True: 
            print(colored('\nParticles detected:\t\t\t', 'green'), sum(abakus_subset.sum(axis=0)), 'pt')
            print(colored('Total particles concentration:\t\t', 'green'), '{:.02f}'.format(ptc_concentration), 'pt/mL\n')
        for i in range(0, (endcolumn-startcolumn)//2): 
            channel_ptc_concentration = abakus_subset[sizes[i]].sum(axis=0)/flow_rate                   # Particles concentration by size [pt/mL] (for each channel)
            ptc_concentration_sizelist.append(np.array([sizes[i], channel_ptc_concentration]))
            if self.terminal==True: print(colored('Particles concentration @', 'green'), sizes[i], colored('\t[mm]:\t', 'green'), channel_ptc_concentration, 'pt/mL')
        if self.terminal==True: print('')

        hist = np.array(abakus_subset.loc[endrow-1])                                                    # Cumulative distribution and plotting options

        return volume, abakus_subset, ptc_concentration, ptc_concentration_sizelist, hist, self.volt_list, self.RAM_list


    # -----------------------------------------------------------------------------------------------------------------------------------------------------#
    # Method which performs the analysis of a single Abakus measurement, retiving informations on total number of counts and sample concentration (gloabal and for each single channel).
    # It os based on the method 'channel_counts' described above.
    # On the left side of the GUI interface, some histograms are generated showing the number of countings for each channel (thus the samople size distribution) and the time
    # evolution of the nnumber of counts as well.

    def scd_analysis(self):

        if self.terminal==True: print(colored('\n\n%-------------------------------------------------------------------------%\n%-------------------------------------------------------------------------%\nSINGLE CUMULATIVE DISTRIBUTION - RESULTS\n%-------------------------------------------------------------------------%\n%-------------------------------------------------------------------------%', 'red'))

        volume, abakus_data, ptc_concentration, ptc_concentration_sizelist, hist, volt_list, RAM_list = self.channel_counts(self.path, self.filename[0], self.skiprows, self.sizes)

        return self.noises, self.measd_flow_rate, volume, abakus_data, ptc_concentration, ptc_concentration_sizelist, hist, volt_list, RAM_list


    # -----------------------------------------------------------------------------------------------------------------------------------------------------#
    # Method to close the existing Abakus instance.

    def close(self):

        try:
            self.first_column_idx = [i for i in range(1, 10000)]
            for i in range(0, len(self.first_column_idx)):
                if i<=34: self.initial_worksheet['A'+str(self.first_column_idx[i])].font = Font(name='Calibri', color=Color('FF0000'), bold=True)
                else: self.initial_worksheet['A'+str(self.first_column_idx[i])].font = Font(name='Calibri', bold=True)
            self.first_column_idx = []
            self.initial_worksheet['B35'].font = Font(name='Calibri', color=Color('FF0000'), bold=True)
            self.initial_worksheet['C35'].font = Font(name='Calibri', color=Color('FF0000'), bold=True)
            self.initial_worksheet['D35'].font = Font(name='Calibri', color=Color('FF0000'), bold=True)
            for i in range(5, 37): 
                self.initial_worksheet.column_dimensions[get_column_letter(i)].width = '20'
                self.initial_worksheet[get_column_letter(i)+'35'].font = Font(name='Calibri', color=Color('FF0000'), bold=True)
        except: print('')

        try:
            self._dev.write(b'C0006\n')                                                                 # Stop the measurement 
            self._dev.write(b'C0000\n')                                                                 # Disconnect Abakus laser sensor
            self._dev.close()                                                                           # Close serial port
        except: print('')


############################################################################################################################################################
############################################################################################################################################################  
